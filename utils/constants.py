import os,sys
from os.path import join as J
"""用于导入项目中不在同一文件夹的库"""
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.config_utils import *

ConstDefaultProfiles = read_init_file()
print(ConstDefaultProfiles)

if "Constants":
    class BoardCst:
        BOARD_CONSTANTS = get_board_default_dict(ConstDefaultProfiles)
        print(BOARD_CONSTANTS)
        ROW_NUM = BOARD_CONSTANTS["row_num"]
        COL_NUM = BOARD_CONSTANTS["col_num"]
        INITIAL_VISION = get_initial_vision_list(BOARD_CONSTANTS)
        RED_INIT_VISION = INITIAL_VISION["red_initial_vision"]
        BLUE_INIT_VISION = INITIAL_VISION["blue_initial_vision"]
        RIVER_BOARD = BOARD_CONSTANTS["border_river"]
        
        if BOARD_CONSTANTS["odd_col_minus_even"] == 1:
            ODD_δx6  = [(0,1),(1,1),(1,0),(0,-1),(-1,0),(-1,1)]
            EVEN_δx6 = [(0,1),(1,0),(1,-1),(0,-1),(-1,-1),(-1,0)]
            ODD_δx12  = [(0,2), (1,2), (2,1), (2,0), (2,-1), (1,-1), 
                        (0,-2), (-1,-1), (-2,-1), (-2,0), (-2,1),(-1,2)]
            EVEN_δx12 = [...]
        
    class CellCst:
        initial_boundaries_type = lambda: {b_index: b_type for b_index, b_type in enumerate([0] * 6, 1)}

        # boundaries_type = {b_index:b_type for b_index,b_type in enumerate([0]*6,1)}
        cell_type = 0
        occupier = 0
        
        class code2name:
            # 单元格边界取值范围字典
            boundary_type_dict = {
            0: "普通边界",
            1: "河流"
            }

            # 单元格类型取值范围字典
            cell_type_dict = {
                0: "道路",
                1: "城市",
                2: "树林",
                3: "山脉",
                4: "荒地"
            }

            # 占领者取值范围字典
            occupier_dict = {
                0: "无人占领",
                1: "红方占领",
                2: "蓝方占领"
            }

if "OtherUtils":
    class DictAttr(dict):
        """一个可以通过属性访问键值的字典类"""

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.__dict__ = self  # 让 __dict__ 直接引用自身

        def __getattr__(self, item):
            try:
                return self[item]
            except KeyError:
                raise AttributeError(f"'{self.__class__.__name__}' object has no attribute '{item}'")

        def __setattr__(self, key, value):
            self[key] = value

        def __delattr__(self, key):
            try:
                del self[key]
            except KeyError:
                raise AttributeError(f"'{self.__class__.__name__}' object has no attribute '{key}'")

        # # 测试
        # data = DictAttr(a=1, b=2, c=3)
        # print(data.a)  # 1
        # print(data['b'])  # 2

        # data.d = 4
        # print(data.d)  # 4
        # print(data['d'])  # 4

        # del data.c
        # print(data)  # {'a': 1, 'b': 2, 'd': 4}

        # # 访问不存在的属性会报错
        # # print(data.x)  # AttributeError: 'DictAttr' object has no attribute 'x'

if "CoordinateUtil":
# class CoordinateUtil:

    # Copyright (c) 2010-2020 openpyxl

    """
    Collection of utilities used within the package and also available for client code
    """
    import re
    from string import digits
    from openpyxl.utils.exceptions import CellCoordinatesException

    # constants
    COORD_RE = re.compile(r'^[$]?([A-Za-z]{1,3})[$]?(\d+)$')
    COL_RANGE = """[A-Z]{1,3}:[A-Z]{1,3}:"""
    ROW_RANGE = r"""\d+:\d+:"""
    RANGE_EXPR = r"""
    [$]?(?P<min_col>[A-Za-z]{1,3})?
    [$]?(?P<min_row>\d+)?
    (:[$]?(?P<max_col>[A-Za-z]{1,3})?
    [$]?(?P<max_row>\d+)?)?
    """
    ABSOLUTE_RE = re.compile('^' + RANGE_EXPR +'$', re.VERBOSE)
    SHEET_TITLE = r"""
    (('(?P<quoted>([^']|'')*)')|(?P<notquoted>[^'^ ^!]*))!"""
    SHEETRANGE_RE = re.compile("""{0}(?P<cells>{1})(?=,?)""".format(
        SHEET_TITLE, RANGE_EXPR), re.VERBOSE)


    def get_column_interval(start, end):
        """
        Given the start and end columns, return all the columns in the series.

        The start and end columns can be either column letters or 1-based
        indexes.
        """
        if isinstance(start, str):
            start = column_index_from_string(start)
        if isinstance(end, str):
            end = column_index_from_string(end)
        return [get_column_letter(x) for x in range(start, end + 1)]



    def coordinate_from_string(coord_string):
        """Convert a coordinate string like 'B12' to a tuple ('B', 12)"""
        match = COORD_RE.match(coord_string)
        if not match:
            msg = f"Invalid cell coordinates ({coord_string})"
            raise CellCoordinatesException(msg)
        column, row = match.groups()
        row = int(row)
        if not row:
            msg = f"There is no row 0 ({coord_string})"
            raise CellCoordinatesException(msg)
        return column, row



    def absolute_coordinate(coord_string):
        """Convert a coordinate to an absolute coordinate string (B12 -> $B$12)"""
        m = ABSOLUTE_RE.match(coord_string)
        if not m:
            raise ValueError(f"{coord_string} is not a valid coordinate range")

        d = m.groupdict('')
        for k, v in d.items():
            if v:
                d[k] = f"${v}"

        if d['max_col'] or d['max_row']:
            fmt = "{min_col}{min_row}:{max_col}{max_row}"
        else:
            fmt = "{min_col}{min_row}"
        return fmt.format(**d)



    def _get_column_letter(col_idx):
        """Convert a column number into a column letter (3 -> 'C')

        Right shift the column col_idx by 26 to find column letters in reverse
        order.  These numbers are 1-based, and can be converted to ASCII
        ordinals by adding 64.

        """
        # these indicies corrospond to A -> ZZZ and include all allowed
        # columns
        if not 1 <= col_idx <= 18278:
            raise ValueError("Invalid column index {0}".format(col_idx))
        letters = []
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx, 26)
            # check for exact division and borrow if needed
            if remainder == 0:
                remainder = 26
                col_idx -= 1
            letters.append(chr(remainder+64))
        return ''.join(reversed(letters))


    _COL_STRING_CACHE = {}
    _STRING_COL_CACHE = {}
    for i in range(1, 18279):
        col = _get_column_letter(i)
        _STRING_COL_CACHE[i] = col
        _COL_STRING_CACHE[col] = i


    def get_column_letter(idx,):
        """Convert a column index into a column letter
        (3 -> 'C')
        """
        try:
            return _STRING_COL_CACHE[idx]
        except KeyError:
            raise ValueError("Invalid column index {0}".format(idx))



    def column_index_from_string(str_col):
        """Convert a column name into a numerical index
        ('A' -> 1)
        """
        # we use a function argument to get indexed name lookup
        try:
            return _COL_STRING_CACHE[str_col.upper()]
        except KeyError:
            raise ValueError("{0} is not a valid column name".format(str_col))



    def range_boundaries(range_string):
        """
        Convert a range string into a tuple of boundaries:
        (min_col, min_row, max_col, max_row)
        Cell coordinates will be converted into a range with the cell at both end
        """
        msg = "{0} is not a valid coordinate or range".format(range_string)
        m = ABSOLUTE_RE.match(range_string)
        if not m:
            raise ValueError(msg)

        min_col, min_row, sep, max_col, max_row = m.groups()

        if sep:
            cols = min_col, max_col
            rows = min_row, max_row

            if not (
                all(cols + rows) or
                all(cols) and not any(rows) or
                all(rows) and not any(cols)
            ):
                raise ValueError(msg)

        if min_col is not None:
            min_col = column_index_from_string(min_col)

        if min_row is not None:
            min_row = int(min_row)

        if max_col is not None:
            max_col = column_index_from_string(max_col)
        else:
            max_col = min_col

        if max_row is not None:
            max_row = int(max_row)
        else:
            max_row = min_row

        return min_col, min_row, max_col, max_row



    def rows_from_range(range_string):
        """
        Get individual addresses for every cell in a range.
        Yields one row at a time.
        """
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        rows = range(min_row, max_row + 1)
        cols = [get_column_letter(col) for col in range(min_col, max_col + 1)]
        for row in rows:
            yield tuple('{0}{1}'.format(col, row) for col in cols)



    def cols_from_range(range_string):
        """
        Get individual addresses for every cell in a range.
        Yields one row at a time.
        """
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        rows = range(min_row, max_row+1)
        cols = (get_column_letter(col) for col in range(min_col, max_col+1))
        for col in cols:
            yield tuple('{0}{1}'.format(col, row) for row in rows)



    def coordinate_to_tuple(coordinate):
        """
        Convert an Excel style coordinate to (row, colum) tuple
        """
        for idx, c in enumerate(coordinate):
            if c in digits:
                break
        col = coordinate[:idx].upper()
        row = coordinate[idx:]
        return int(row), _COL_STRING_CACHE[col]



    def range_to_tuple(range_string):
        """
        Convert a worksheet range to the sheetname and maximum and minimum
        coordinate indices
        """
        m = SHEETRANGE_RE.match(range_string)
        if m is None:
            raise ValueError("Value must be of the form sheetname!A1:E4")
        sheetname = m.group("quoted") or m.group("notquoted")
        cells = m.group("cells")
        boundaries = range_boundaries(cells)
        return sheetname, boundaries



    def quote_sheetname(sheetname):
        """
        Add quotes around sheetnames if they contain spaces.
        """
        if "'" in sheetname:
            sheetname = sheetname.replace("'", "''")

        sheetname = u"'{0}'".format(sheetname)
        return sheetname
